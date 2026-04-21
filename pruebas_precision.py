"""
pruebas_precision.py
Evalúa la precisión del Knowledge Base con 100 preguntas de referencia sobre el reporte 0430.
Las respuestas se guardan en un Excel para revisión manual.
"""

import os
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

API_URL = "http://localhost:8000"

# ══════════════════════════════════════════════════════════
# 100 PREGUNTAS DE REFERENCIA
# Mezcla de simples, intermedias y complicadas
# Formato: (pregunta, comando, reporte, categoria)
# ══════════════════════════════════════════════════════════

PREGUNTAS = [
    # ── SIMPLES — campos básicos ──────────────────────────
    ("¿Qué es el campo RFC del acreditado?",                                       "campo",    "0430", "Simple"),
    ("¿Qué es el campo CURP del acreditado?",                                      "campo",    "0430", "Simple"),
    ("¿Qué es el campo Nombre del acreditado?",                                    "campo",    "0430", "Simple"),
    ("¿Qué es el campo Periodo?",                                                  "campo",    "0430", "Simple"),
    ("¿Qué es el campo Reporte?",                                                  "campo",    "0430", "Simple"),
    ("¿Qué es el campo Tipo de cartera?",                                          "campo",    "0430", "Simple"),
    ("¿Qué es el campo Tipo de anexo?",                                            "campo",    "0430", "Simple"),
    ("¿Qué es el campo Destino del crédito?",                                      "campo",    "0430", "Simple"),
    ("¿Qué es el campo Moneda de la línea de crédito?",                            "campo",    "0430", "Simple"),
    ("¿Qué es el campo Nacionalidad del acreditado?",                              "campo",    "0430", "Simple"),
    ("¿Qué es el campo Actividad económica del acreditado?",                       "campo",    "0430", "Simple"),
    ("¿Qué es el campo Grupo de riesgo?",                                          "campo",    "0430", "Simple"),
    ("¿Qué es el campo Acreditado relacionado?",                                   "campo",    "0430", "Simple"),
    ("¿Qué es el campo Clave LEI?",                                                "campo",    "0430", "Simple"),
    ("¿Qué es el campo Tipo de alta del crédito?",                                 "campo",    "0430", "Simple"),
    ("¿Qué es el campo Tipo de tasa de interés?",                                  "campo",    "0430", "Simple"),
    ("¿Qué es el campo Periodicidad de pagos de capital?",                         "campo",    "0430", "Simple"),
    ("¿Qué es el campo Periodicidad de pagos de intereses?",                       "campo",    "0430", "Simple"),
    ("¿Qué es el campo Fecha de otorgamiento de la línea de crédito?",             "campo",    "0430", "Simple"),
    ("¿Qué es el campo Fecha de vencimiento de la línea de crédito?",              "campo",    "0430", "Simple"),

    # ── SIMPLES — cálculo de campos calculados ─────────────
    ("¿Qué campos se calculan en el reporte 0430?",                                "consulta", "0430", "Simple"),
    ("¿Cómo se calcula el campo 20?",                                              "calculo",  "0430", "Simple"),
    ("¿Cómo se calcula el campo 23?",                                              "calculo",  "0430", "Simple"),
    ("¿Cómo se calcula el campo 51?",                                              "calculo",  "0430", "Simple"),
    ("¿Cómo se calcula el campo 52?",                                              "calculo",  "0430", "Simple"),

    # ── SIMPLES — origen de campos ─────────────────────────
    ("¿El campo RFC es de captura manual o calculado?",                            "campo",    "0430", "Simple"),
    ("¿El campo Municipio del domicilio es de catálogo o manual?",                 "campo",    "0430", "Simple"),
    ("¿El campo Clave de la institución es de catálogo?",                          "campo",    "0430", "Simple"),
    ("¿El campo Nombre del acreditado viene de PERSONA o LINEA_CREDITO?",          "campo",    "0430", "Simple"),
    ("¿El campo Fecha de otorgamiento viene de la persona o de la línea?",         "campo",    "0430", "Simple"),

    # ── INTERMEDIAS — validaciones específicas ─────────────
    ("¿Cuántos caracteres debe tener el RFC del acreditado?",                      "consulta", "0430", "Intermedia"),
    ("¿Qué formato debe tener el periodo en el reporte 0430?",                     "consulta", "0430", "Intermedia"),
    ("¿Qué catálogo se usa para el campo Destino del crédito?",                    "consulta", "0430", "Intermedia"),
    ("¿Qué valores válidos tiene el campo Tipo de cartera?",                       "consulta", "0430", "Intermedia"),
    ("¿Qué valores válidos tiene el campo Tipo de anexo?",                         "consulta", "0430", "Intermedia"),
    ("¿Qué valores válidos tiene el campo Tipo de alta del crédito?",              "consulta", "0430", "Intermedia"),
    ("¿Qué valores válidos tiene el campo Línea revocable o irrevocable?",         "consulta", "0430", "Intermedia"),
    ("¿Qué valores válidos tiene el campo Prelación de pago?",                     "consulta", "0430", "Intermedia"),
    ("¿Qué catálogo usa el campo Nacionalidad del acreditado?",                    "consulta", "0430", "Intermedia"),
    ("¿Qué catálogo usa el campo Actividad económica del acreditado?",             "consulta", "0430", "Intermedia"),

    # ── INTERMEDIAS — fórmulas y reglas ───────────────────
    ("¿Cómo se construye el Identificador del Crédito Metodología CNBV?",          "calculo",  "0430", "Intermedia"),
    ("¿Qué datos se necesitan para calcular el campo 20?",                         "calculo",  "0430", "Intermedia"),
    ("¿Cómo se extrae el municipio de destino del crédito?",                       "calculo",  "0430", "Intermedia"),
    ("¿Cómo se convierte el monto de la línea de crédito a pesos?",                "calculo",  "0430", "Intermedia"),
    ("¿Qué es el Anexo 3 y para qué sirve en el reporte 0430?",                    "consulta", "0430", "Intermedia"),
    ("¿Qué diferencia hay entre el campo 19 y el campo 20?",                       "consulta", "0430", "Intermedia"),
    ("¿Qué diferencia hay entre el campo 23 y el campo 24?",                       "consulta", "0430", "Intermedia"),
    ("¿Qué diferencia hay entre el campo 51 y el campo 50?",                       "consulta", "0430", "Intermedia"),
    ("¿Qué diferencia hay entre el campo 9 y el campo 51?",                        "consulta", "0430", "Intermedia"),
    ("¿Cuál es la diferencia entre Monto autorizado (campo 48) y Monto valorizado (campo 23)?", "consulta", "0430", "Intermedia"),

    # ── INTERMEDIAS — reglas por tipo de cartera ──────────
    ("¿Cuándo el RFC del acreditado debe iniciar con guion bajo?",                 "consulta", "0430", "Intermedia"),
    ("¿Qué tipo de carteras requieren que el RFC inicie con guion bajo?",          "consulta", "0430", "Intermedia"),
    ("¿Cuándo aplica el campo CURP del acreditado?",                               "consulta", "0430", "Intermedia"),
    ("¿Qué campos son obligatorios para personas morales?",                        "consulta", "0430", "Intermedia"),
    ("¿Qué campos son obligatorios para personas físicas?",                        "consulta", "0430", "Intermedia"),
    ("¿Cuándo el Nombre del acreditado no debe incluir tipo de sociedad?",         "consulta", "0430", "Intermedia"),
    ("¿Qué condiciones aplican para el campo Grupo de riesgo?",                    "consulta", "0430", "Intermedia"),
    ("¿Cuándo aplica el campo Clave LEI?",                                         "consulta", "0430", "Intermedia"),
    ("¿Qué validaciones aplican al campo Fecha de otorgamiento?",                  "consulta", "0430", "Intermedia"),
    ("¿Cuándo es obligatorio el campo Número de consulta a la SIC?",               "consulta", "0430", "Intermedia"),

    # ── INTERMEDIAS — secciones del reporte ───────────────
    ("¿Cuántas secciones tiene el reporte 0430?",                                  "consulta", "0430", "Intermedia"),
    ("¿Qué campos pertenecen a la sección de identificador del reporte?",          "consulta", "0430", "Intermedia"),
    ("¿Qué campos pertenecen a la sección de identificador del acreditado?",       "consulta", "0430", "Intermedia"),
    ("¿Qué campos pertenecen a la sección de identificador del crédito?",          "consulta", "0430", "Intermedia"),
    ("¿Qué campos pertenecen a la sección de condiciones del crédito?",            "consulta", "0430", "Intermedia"),
    ("¿Qué es el SITI y cómo se relaciona con el reporte 0430?",                  "consulta", "0430", "Intermedia"),
    ("¿Qué es el CUB y cómo aplica al reporte 0430?",                             "consulta", "0430", "Intermedia"),
    ("¿Cuál es el objetivo del reporte 0430?",                                     "consulta", "0430", "Intermedia"),
    ("¿Quiénes están obligados a enviar el reporte 0430?",                         "consulta", "0430", "Intermedia"),
    ("¿Con qué frecuencia se envía el reporte 0430?",                              "consulta", "0430", "Intermedia"),

    # ── COMPLICADAS — reglas cruzadas ─────────────────────
    ("¿Qué validaciones aplican al Identificador del Acreditado cuando el mismo acreditado aparece en múltiples registros?", "consulta", "0430", "Complicada"),
    ("¿Qué consistencia debe haber entre el RFC, Nombre y CURP del mismo acreditado en diferentes registros?", "consulta", "0430", "Complicada"),
    ("¿Cuándo el campo Localidad del domicilio del acreditado determina el Municipio y Estado?", "consulta", "0430", "Complicada"),
    ("¿Cómo se relacionan los campos 8, 9 y 10 entre sí?",                        "consulta", "0430", "Complicada"),
    ("¿Cómo se relacionan los campos 50, 51 y 52 entre sí?",                      "consulta", "0430", "Complicada"),
    ("¿Qué ocurre con el campo Identificador del Acreditado cuando el Tipo de Cartera es 299?", "consulta", "0430", "Complicada"),
    ("¿Cuándo el campo Nombre del acreditado debe coincidir con la Base de Fideicomisos de la CNBV?", "consulta", "0430", "Complicada"),
    ("¿Qué condiciones determinan si una línea de crédito es revocable o irrevocable?", "consulta", "0430", "Complicada"),
    ("¿Cómo se calcula el CAT del campo 47 y qué componentes incluye?",            "calculo",  "0430", "Complicada"),
    ("¿Qué relación hay entre el campo Tipo de Alta (33) y el momento del reporte?", "consulta", "0430", "Complicada"),

    # ── COMPLICADAS — metodología CNBV ────────────────────
    ("¿Qué es la metodología CNBV para el Identificador del Crédito y en qué Anexo se describe?", "consulta", "0430", "Complicada"),
    ("¿Qué componentes forman el Identificador del Crédito del campo 20?",         "calculo",  "0430", "Complicada"),
    ("¿Cómo se determina el Tipo de Crédito para el Identificador del campo 20?",  "consulta", "0430", "Complicada"),
    ("¿Qué pasa si el Identificador del Crédito del campo 20 ya existe en el reporte de seguimiento?", "consulta", "0430", "Complicada"),
    ("¿Qué relación hay entre el campo 20 del 0430 y el reporte 0431?",            "consulta", "0430", "Complicada"),
    ("¿Cuándo aplica el campo Identificador Crédito Línea Grupal (campo 21)?",     "consulta", "0430", "Complicada"),
    ("¿Cómo se valida que el Porcentaje de Participaciones Federales (campo 31) sea congruente con el Tipo de Cartera?", "consulta", "0430", "Complicada"),
    ("¿Qué restricciones tiene el campo Destino del Crédito según el Tipo de Cartera?", "consulta", "0430", "Complicada"),
    ("¿Cuándo el campo Monto de la Línea Autorizado en Pesos (campo 23) difiere del campo 24?", "consulta", "0430", "Complicada"),
    ("¿Qué condiciones hacen que el campo Fecha Máxima para Disponer (campo 27) sea obligatorio?", "consulta", "0430", "Complicada"),

    # ── COMPLICADAS — casos edge ───────────────────────────
    ("¿Qué diferencia hay entre un crédito de Tipo de Alta 1 y Tipo de Alta 2?",  "consulta", "0430", "Complicada"),
    ("¿Qué validaciones especiales aplican cuando el Tipo de Cartera es 141 o 142?", "consulta", "0430", "Complicada"),
    ("¿Cómo se reporta un crédito en moneda extranjera en el campo 23?",           "consulta", "0430", "Complicada"),
    ("¿Qué pasa con el campo Diferencial sobre Tasa de Referencia si la tasa es fija?", "consulta", "0430", "Complicada"),
    ("¿Cuándo el campo Número de meses de gracia para amortizar capital puede ser cero?", "consulta", "0430", "Complicada"),
    ("¿Qué validaciones aplican al campo Clave de la Institución Otorgante (campo 32)?", "consulta", "0430", "Complicada"),
    ("¿Cuándo es obligatorio reportar las comisiones de apertura (campos 43 y 44)?", "consulta", "0430", "Complicada"),
    ("¿Qué relación hay entre el campo Porcentaje de Participaciones Federales y el Destino del Crédito?", "consulta", "0430", "Complicada"),
    ("¿Qué ocurre si el acreditado registrado en el 0430 no aparece en el reporte de Probabilidad de Incumplimiento?", "consulta", "0430", "Complicada"),
    ("¿Cómo se determina la Actividad Económica a la que se destinará el crédito (campo 53) cuando es diferente a la del acreditado (campo 12)?", "consulta", "0430", "Complicada"),
]

# ══════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════

def preguntar(pregunta, cmd, reporte):
    """Manda la pregunta al Knowledge Base y devuelve la respuesta"""
    try:
        url = f"{API_URL}/{cmd}"
        body = {"pregunta": pregunta, "reporte": reporte, "session_id": "prueba_precision"}
        r = requests.post(url, json=body, timeout=120)
        if r.status_code == 200:
            return r.json().get("respuesta", "Sin respuesta")
        return f"Error {r.status_code}: {r.text[:100]}"
    except Exception as e:
        return f"Error: {str(e)[:100]}"


def generar_excel(resultados, archivo="resultados_precision.xlsx"):
    """Genera Excel con preguntas, respuestas y columna para calificación manual"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def font(bold=False, color='FF000000', size=11):
        return Font(bold=bold, color=color, size=size, name='Arial')

    def align(h='left', wrap=True):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

    # Columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10

    # Header
    headers = ['#', 'Categoría', 'Pregunta', 'Comando', 'Respuesta del sistema', 'Calificación', 'Notas']
    header_fill = fill('FF2E75B6')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.font = font(bold=True, color='FFFFFFFF')
        c.fill = header_fill
        c.alignment = align(h='center')

    # Colores por categoría
    cat_colors = {
        'Simple':     'FFE2EFDA',
        'Intermedia': 'FFFFF2CC',
        'Complicada': 'FFFCE4D6',
    }

    # Datos
    for i, (num, categoria, pregunta, cmd, reporte, respuesta) in enumerate(resultados, 2):
        color = cat_colors.get(categoria, 'FFFFFFFF')
        ws.row_dimensions[i].height = 80

        vals = [num, categoria, pregunta, cmd, respuesta, '', '']
        aligns_row = ['center', 'center', 'left', 'center', 'left', 'center', 'left']

        for col, (val, aln) in enumerate(zip(vals, aligns_row), 1):
            c = ws.cell(row=i, column=col)
            c.value = val
            c.font = font(size=10)
            c.fill = fill(color)
            c.alignment = align(h=aln)

    # Instrucciones en columna F
    ws['F1'].value = 'Calificación'
    instruccion = ws.cell(row=2, column=7)
    instruccion.value = 'Usa: ✅ Correcta | ❌ Incorrecta | ⚠️ Parcial'

    # Totales al final
    total_row = len(resultados) + 3
    ws.cell(row=total_row, column=2).value = 'Total preguntas:'
    ws.cell(row=total_row, column=3).value = len(resultados)
    ws.cell(row=total_row + 1, column=2).value = 'Correctas (✅):'
    ws.cell(row=total_row + 1, column=3).value = '=COUNTIF(F2:F101,"✅")'
    ws.cell(row=total_row + 2, column=2).value = 'Incorrectas (❌):'
    ws.cell(row=total_row + 2, column=3).value = '=COUNTIF(F2:F101,"❌")'
    ws.cell(row=total_row + 3, column=2).value = 'Parciales (⚠️):'
    ws.cell(row=total_row + 3, column=3).value = '=COUNTIF(F2:F101,"⚠️")'
    ws.cell(row=total_row + 4, column=2).value = 'Precisión:'
    ws.cell(row=total_row + 4, column=3).value = f'=F{total_row+1}/F{total_row}'

    wb.save(archivo)
    print(f"\n✅ Excel guardado: {archivo}")


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("🔍 Pruebas de Precisión — Knowledge Base CNBV")
    print("=" * 60)
    print(f"Total preguntas: {len(PREGUNTAS)}")
    print("Asegúrate de que el backend esté corriendo en :8000\n")

    # Crear sesión de prueba
    try:
        requests.post(f"{API_URL}/sesiones", json={"nombre": "prueba_precision", "usuario": "admin"})
    except:
        pass

    resultados = []
    errores = 0

    for i, (pregunta, cmd, reporte, categoria) in enumerate(PREGUNTAS, 1):
        print(f"[{i:03d}/{len(PREGUNTAS)}] {categoria:12} | {pregunta[:60]}...")
        respuesta = preguntar(pregunta, cmd, reporte)

        if respuesta.startswith("Error"):
            errores += 1
            print(f"         ❌ {respuesta}")
        else:
            print(f"         ✅ {respuesta[:80]}...")

        resultados.append((i, categoria, pregunta, cmd, reporte, respuesta))
        time.sleep(0.5)  # pequeña pausa entre preguntas

    print(f"\n{'='*60}")
    print(f"Completado: {len(PREGUNTAS)} preguntas | {errores} errores de conexión")
    generar_excel(resultados)
    print("\nAbre resultados_precision.xlsx y califica cada respuesta:")
    print("  ✅ Correcta  — la respuesta es precisa y completa")
    print("  ❌ Incorrecta — la respuesta es incorrecta o alucina datos")
    print("  ⚠️ Parcial   — la respuesta es correcta pero incompleta")